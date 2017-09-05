#!/usr/bin/env python

# given a plot output directory, produces a spreadsheet with the fit parameters

import sys, os, re
import GrandUnificationPlot
import utils

#----------------------------------------------------------------------

def getAverageNumVerticesFromTasks(tasks):
    from NumVertices import NumVertices

    for task in tasks:
        if isinstance(task, NumVertices):
            return task.avg_num_vertices

    # not found
    return None

#----------------------------------------------------------------------

from openpyxl.utils import _get_column_letter

def coordToName(row, col, rowPrefix = "", colPrefix = ""):
    # row and col are one-based

    return colPrefix + _get_column_letter(col) + rowPrefix + str(row)

#----------------------------------------------------------------------


class SingleGroupSheet:
    # fills a single spreadsheet for a given group

    #----------------------------------------
    
    def __init__(self, workbook, groupingName, evolutionData, avgNumVertices, 
                 triggerRateKHz, sheetName = None):

        # @param sheetName if not None will override the name of the worksheet
        #        (which is otherwise taken from groupingName)

        assert len(evolutionData) >= 1

        self.wb = workbook
        self.groupingName = groupingName
        self.evolutionData = evolutionData
        self.avgNumVertices = avgNumVertices
        self.triggerRateKHz = triggerRateKHz
        self.sheetName = sheetName

        # number of coefficients for each fit
        self.numCoeffs = len(evolutionData[0]['coeffs'])

    #----------------------------------------

    def __makeHeaderCells(self):
        # write average number of vertices
        # and trigger rate
        #
        # @return the next row index to be used

        self.avgNumVtxCell = (1,2)
        self.avgNumVtxCellName = coordToName(*self.avgNumVtxCell, rowPrefix = '$') # B$1

        self[(self.avgNumVtxCell[0], self.avgNumVtxCell[1] - 1)] = 'avg. number of vertices'
        if self.avgNumVertices == None:
            self[self.avgNumVtxCell] = "unknown"
        else:
            self.makeNumericCell(self.avgNumVtxCell, self.avgNumVertices, "0.0")

        self.ws.column_dimensions[_get_column_letter(self.avgNumVtxCell[1] - 1)].width = 20

        #----------
        # trigger rate
        #----------
        self.triggerRateCell = (1,9)
        self.triggerRateCellName = coordToName(*self.triggerRateCell, rowPrefix = "$", colPrefix = "$")

        self[(self.triggerRateCell[0], self.triggerRateCell[1] - 1)] = 'trigger rate [kHz]:' # H1
        self.ws.column_dimensions[_get_column_letter(self.triggerRateCell[1] - 1)].width = 14
        self[self.triggerRateCell] = str(self.triggerRateKHz) # 'I1'

        return 3

    #----------------------------------------

    # helper function to assign to the current sheet
    # by row and column number
    def __setitem__(self, key, value):
        self.ws._get_cell(*key).value = value

    #----------------------------------------

    def __fillInputData(self, topLeft):
        # fills the fitted data per FED group on which almost
        # any other numbers are based on

        firstRow, firstCol = topLeft

        #----------
        # column headers
        #----------
        self[(firstRow, 4)] = 'sum data sizes [kByte/ev]' # D3
        # self.ws['F3'] = '=CONCATENATE("data size at ";CELL("contents";A2);" vertices") '
        # self.ws['F3'] = '=\"data size at \"\&A2\&\" vertices\"'

        # does not work with POI
        # self.ws['F3'] = '="data size at "&A2&" vertices"'

        self[(firstRow + 2, firstCol)]     = 'FED group'               # A5                  
        self[(firstRow + 2, firstCol + 1)] = 'number of feds'          # B5
        self[(firstRow + 2, firstCol + 2)] = 'subsys' # do we still need this ? # C5

        for power in range(self.numCoeffs):
            self[(firstRow + 2, firstCol + 3 + power)] = utils.getPowerName(power)  # D5

        self.ws.column_dimensions[_get_column_letter(firstCol + 1)].width = 14

        #----------
        # fill the evolution data
        #----------

        for rowOffset, data in enumerate(self.evolutionData):
            thisRow = rowOffset + firstRow + 3
            self[(thisRow, firstCol)]     = data['subsystem']                          # A%d
            self[(thisRow, firstCol + 1)] = data['numFeds']                            # B%d

            coeffs = data['coeffs']

            for power in range(self.numCoeffs):
                self.makeNumericCell((thisRow, firstCol + 3 + power), data['coeffs'][power], "#,##0.000") # D%d

    #----------------------------------------

    def __fillDataSizeAtNumVertices(self, topLeft, topLeftInputData, numVtxCellName):
        # produces cells calculating the data size at a given number
        # of vertices
        #
        # @param topLeftInputData is (row,col) of the first data cell
        # (offset) of the input data

        firstRow, firstCol = topLeft

        firstRowInputData, inputCol = topLeftInputData

        #----------
        # title cells
        #----------

        self[(firstRow,     firstCol)] = 'data size [kByte/ev]'                                            # G3
        self[(firstRow + 1, firstCol)] = '=CONCATENATE("at ",TEXT(%s,"0.0")," vertices")' % numVtxCellName # G4

        #----------
        # fill the formulas
        #----------

        for i in range(len(self.evolutionData)):
            thisRow = firstRow + 3 + i
            # need absolute cell rows to allow sorting by the user

            inputRow = firstRowInputData + i 
            
            parts = []

            for power in range(self.numCoeffs):
                thisPart = coordToName(inputRow, inputCol + power) # D%d

                if power == 1:
                    thisPart += " * " + numVtxCellName # B$1
                elif power >= 2:
                    thisPart += " * POWER(%s, %d)" % (numVtxCellName, power)

                parts.append(thisPart)

            # G%d =D%d+E%d*B$1
            self.makeNumericCell((thisRow, firstCol), 
                                 "=" + " + " .join(parts), 
                                 "#,##0.000")


    #----------------------------------------

    def __fillDataRateOffsetSlope(self, topLeft, topLeftInputData, triggerRateCellName,
                                  divideByNumFeds, topLeftNumFeds = None):
        # produces columns with data rate offset and slope

        if divideByNumFeds:
            assert topLeftNumFeds != None

        firstRow, firstCol = topLeft

        firstRowInputData, firstColInputData = topLeftInputData

        #----------
        # titles
        #----------

        if divideByNumFeds:
            self[(firstRow,    firstCol)] = "per FED data rate [MByte/s]" # L3
        else:
            self[(firstRow,    firstCol)] = "data rate [MByte/s]" # I3

        self[(firstRow + 1,firstCol)]     = '=CONCATENATE("at ", TEXT(%s,"0.0")," kHz trigger rate")' % triggerRateCellName # I4

        for power in range(self.numCoeffs):
            self[(firstRow + 2, firstCol + power)] = utils.getPowerName(power)  # I5

        #----------
        # equations
        #----------

        for i in range(len(self.evolutionData)):
            thisRow = firstRow + 3 + i

            inputRow = firstRowInputData + i

            # loop over offset and slope
            for power in range(self.numCoeffs):

                if divideByNumFeds:
                    self.makeNumericCell((thisRow, firstCol + power),  "=%s*%s/%s" % ( # I%d =D%d*I$1/B%d
                            coordToName(inputRow, firstColInputData + power), # D%d
                            triggerRateCellName,                                # I$1
                            coordToName(topLeftNumFeds[0] + i, topLeftNumFeds[1]), # B%d (number of FEDs in this group)
                            ),  
                                 "#,##0.000") 

                else:
                    self.makeNumericCell((thisRow, firstCol + power),  "=%s*%s" % ( # I%d =D%d*I$1
                            coordToName(inputRow, firstColInputData + power), # D%d
                            triggerRateCellName),    # I$1
                                 "#,##0.000") 
            # loop over coefficients of polynomial

    #----------------------------------------

    def __fillLimit(self, topLeft, topLeftInputData,
                    maxDataRate,
                    maxDataRateCell,
                    triggerRateCellName,
                    divideByNumFeds, topLeftNumFeds = None,
                    usePileup = False,
                    ):
        
        # @param usePileup: if this is True, the limit is shown in terms of pileup (number of interactions
        #                   per bunch crossing) rather than number of vertices, assuming
        #                   one interaction leads to 0.7 vertices on average
        # 
        # @return the number of solution columns added

        assert self.numCoeffs <= 3, "only polynomials up to degree two are supported"
        
        if divideByNumFeds:
            assert topLeftNumFeds != None

        limitColumn = topLeft[1]

        maxDataRateCellName = coordToName(*maxDataRateCell, rowPrefix = "$")

        #----------
        # determine group title
        #----------
        if divideByNumFeds:
            title = "per FED data rate limit"
        else:
            title = "group"
            if self.groupingName != None:
                mo = re.match("by (.*)$", self.groupingName)
                if mo:
                    title = mo.group(1)

            title = title + " data rate limit"

        self[(maxDataRateCell[0], maxDataRateCell[1] - 1)] = "per " + title + " [MByte/s]"
        self.ws.column_dimensions[_get_column_letter(maxDataRateCell[1])].width = 27

        self[maxDataRateCell]  = str(maxDataRate) # P1
        
        #----------

        if usePileup:
            self[topLeft] = "pileup for " + title # O3
        else:
            self[topLeft] = "# vertices for " + title # O3


        # number of columns added
        retval = 1

        # number of vertices where we cross the limit is
        # (data rate limit - data rate offset) / data rate slope
        #
        # where data rates are (fed size) * (trigger rate)

        # e.g. =(P$1-D7*I$1/B7)/(E7*I$1/B7)
        # 
        # where I$1 is the trigger rate cell
        # and   P$1 is the maximum data rate cell
        #       D7  is the fragment/group size offset
        #       E7  is the fragment/group size slope
        #       B7  is the number of feds in this group
            
        # if we should display the pileup instead, we divide this by 0.7

        for i in range(len(self.evolutionData)):
            thisRow = topLeft[0] + 3 + i

            inputRow = topLeftInputData[0] + i

            # solution for linear models:
            # 
            #   numVertices = (limit - data rate const) / data rate slope
            #
            # solution for quadratic models:
            #
            # (-b +/- sqrt( b^2 - 4ac) ) / 2a
            # 
            # where 
            #   a = quadratic term
            #   b = linear term
            #   c = const term - limit
            #
            # for lack of a heuristic which of the two solutions is the
            # wanted one (b can be negative) we just put two columns for the moment

            if self.numCoeffs < 2:
                # no dependence on pileup
                self[(thisRow, limitColumn)] = 'N/A'
                continue

            # build expressions for the different coefficients, divided by number
            # or FEDS or not
            coeffExprs = []
            for power in range(self.numCoeffs):

                coeffCell = coordToName(inputRow, topLeftInputData[1] + power) # D%d

                if divideByNumFeds:
                    numFedsCell = coordToName(topLeftNumFeds[0] + i, topLeftNumFeds[1]) # B%d 
                    coeffExprs.append("%s * %s / %s"  % (coeffCell, triggerRateCellName, numFedsCell))
                else:
                    coeffExprs.append("%s * %s"  % (coeffCell, triggerRateCellName))

            if self.numCoeffs == 2:
                # linear model
                exprs = [ "({maxDataRate} - ({coeff0})) / ({coeff1)".format(
                        maxDataRate = maxDataRateCellName,      # P$1
                        coeff0 = coeffExprs[0],
                        coeff1 = coeffExprs[0],
                        )
                          ]

                retval = 1
            elif self.numCoeffs == 3:
                # quadratic model

                replacements = dict(
                    a = coeffExprs[2],
                    b = coeffExprs[1],
                    c = "(%s - %s)" % (coeffExprs[0], maxDataRateCellName)
                    )

                # add a protection against negative values under the square root
                determinantExpr = "POWER({b},2) - 4 * ({a}) * ({c})"

                exprs = []

                for sign in ('-', '+'):
                
                    # solution of the quadratic equation
                    solutionExpr = "(-({b}) " + sign + " SQRT(" + determinantExpr + ")) / (2 * ({a}))"

                    if usePileup:
                        solutionExpr = "(%s) / 0.7" % solutionExpr

                    exprs.append(
                        ("IF(" + determinantExpr + ">=0," + 

                         # determinant non-negative
                         solutionExpr + 
                         
                         "," +
                         '"(no solution)"' + # determinant negative
                         ")"
                         ).format(**replacements)
                        )
                # end of loop over signs

                retval = 2

            else:
                raise Exception("uncaught case - internal error")

            for colOffset, expr in enumerate(exprs):
                self.makeNumericCell((thisRow, limitColumn + colOffset), # O%d
                                     "=" + expr,
                                     "0.0")
            # loop over expressions
        # end of loop over rows

        # return number of columns added
        return retval

    #----------------------------------------

    def __fillUncertainties(self, topLeftUncertainties):

        firstRow, firstCol = topLeftUncertainties

        #----------
        # titles
        #----------

        self[(firstRow,     firstCol)]     = 'one sigma spread on data sizes [kByte/ev]' # Q3 

        for power in range(self.numCoeffs):
            self[(firstRow + 2, firstCol + power)] = utils.getPowerName(power) # Q5

        #----------
        # data
        #----------

        for i, data in enumerate(self.evolutionData):
            thisRow = firstRow + 3 + i 

            for power in range(self.numCoeffs):
                self.makeNumericCell((thisRow, firstCol + power), data['uncertCoeffs'][power], "#,##0.000") # Q%d

    #----------------------------------------

    def __createWorkSheet(self):
        self.ws = self.wb.create_sheet()

        if self.sheetName != None:
            self.ws.title = self.sheetName
        else:
            # determine from groupingName (which is also used 
            # to determine other quantities)
        
            if self.groupingName == "" or self.groupingName == None:
                self.ws.title = "-"
            else:
                self.ws.title = self.groupingName

    #----------------------------------------

    def fillSheet(self):
        
        # create a new worksheet
        self.__createWorkSheet()

        #----------
        # average number of vertices and
        # nominal trigger rate
        #----------
        row = self.__makeHeaderCells()

        # fill input data (fit results)
        self.__fillInputData((row, 1))

        topLeftInputData = (row + 3, 4)
        topLeftNumFeds   = (row + 3, 2)

        nextCol = 5 + self.numCoeffs

        #----------
        # data size at given number of vertices
        #----------
        self.__fillDataSizeAtNumVertices(topLeft = (row, nextCol), 
                                         topLeftInputData = topLeftInputData,
                                         numVtxCellName = self.avgNumVtxCellName)

        nextCol += 2

        #--------------------
        # data rate
        #--------------------
        self.__fillDataRateOffsetSlope(topLeft = (row, nextCol),
                                       topLeftInputData = topLeftInputData,
                                       triggerRateCellName = self.triggerRateCellName,
                                       divideByNumFeds = False)

        nextCol += 1 + self.numCoeffs

        #----------
        # per FED data rate
        #----------

        self.__fillDataRateOffsetSlope(topLeft = (row, nextCol),
                                       topLeftInputData = topLeftInputData,
                                       triggerRateCellName = self.triggerRateCellName,
                                       divideByNumFeds = True,
                                       topLeftNumFeds = topLeftNumFeds)

        nextCol += 1 + self.numCoeffs
        #----------
        # when do we reach 200 MByte/s per FED ?
        #----------

        if self.numCoeffs <= 3:
            maxDataRateCell = (1, nextCol)

            # maximum data rate per FED
            if self.groupingName == 'by fedbuilder':
                maxDataRate = 4000 # in MByte/s
                divideByNumFeds = False
            else:
                maxDataRate = 200 # in MByte/s
                divideByNumFeds = True

            # now this is independent on any precalculations
            numCols = self.__fillLimit(topLeft = (row, nextCol),  # column O
                                       topLeftInputData = topLeftInputData,
                                       maxDataRate = maxDataRate,
                                       maxDataRateCell = maxDataRateCell,
                                       triggerRateCellName = self.triggerRateCellName,
                                       divideByNumFeds = divideByNumFeds,
                                       topLeftNumFeds = topLeftNumFeds
                                       )

            nextCol += 1 + numCols

        #----------
        # offset and slope of uncertainties fit
        #----------

        topLeftUncertainties = (row, nextCol)
        self.__fillUncertainties(topLeftUncertainties)

    #----------------------------------------

    def __conditionalFormattingForDataRate(self, topLeft, maxDataRateCell):

        # applies conditional formatting to absolute data rate cells
        # for cells exceeding the data rate limit

        firstRow, firstCol = topLeft

        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles.fonts import Font
        from openpyxl.styles.colors import Color
        from openpyxl.styles.fills import PatternFill

        # red font and fill
        font = Font(color = Color(rgb='FF9C0006'))
        fill = PatternFill(bgColor = Color(rgb='FFFFC7CE'))

        # range where this formatting should be applied to
        theRange = '%s:%s' % (coordToName(firstRow + 3, firstCol),
                              coordToName(firstRow + 3 + len(self.evolutionData) - 1, firstCol)
                              )


        # see e.g. http://stackoverflow.com/a/31326228/288875
        # 
        # it looks like conditional formatting on multiple columns does not work
        # properly, the formatting for the first column was taken from the last column etc.
        for i in range(len(self.evolutionData)):

            thisCell = coordToName(firstRow + 3 + i, firstCol)

            self.ws.conditional_formatting.add(
                thisCell
                ,
                FormulaRule(formula = ['%s - %s > 0' % (thisCell, coordToName(*maxDataRateCell))],

                        font = font,
                        fill = fill,
                        )
            )




    #----------------------------------------

    def __fillDataRateAtPileup(self,
                               topLeft,
                               topLeftInputData,
                               triggerRateCellName,
                               pileupCellName,
                               ):
        # fills cells (single column, not offset and slope like __fillDataRateOffsetSlope(..)) 
        # with extrapolated data rates for a given pileup value

        firstRow, firstCol = topLeft

        firstRowInputData, firstColInputData = topLeftInputData

        #----------
        # titles
        #----------

        self[(firstRow,    firstCol)] = "data rate [MByte/s]" 
        self[(firstRow + 1,firstCol)] = '=CONCATENATE("at ", TEXT(%s,"0.0")," pileup")' % pileupCellName 
        self[(firstRow + 2,firstCol)] = '=CONCATENATE("and ", TEXT(%s,"0.0")," kHz trigger rate")' % triggerRateCellName 

        #----------
        # equations
        #----------

        for i in range(len(self.evolutionData)):
            thisRow = firstRow + 3 + i

            inputRow = firstRowInputData + i

            parts = []

            for power in range(self.numCoeffs):
                thisPart = coordToName(inputRow, topLeftInputData[1] + power) # D%d

                if power == 1:
                    thisPart += " * %s * 0.7 " % pileupCellName # B$1
                elif power >= 2:
                    thisPart += " * POWER(%s * 0.7, %d)" % (pileupCellName, power)

                parts.append(thisPart)

            expr = "(%s) * %s" % (" + ".join(parts),
                                  triggerRateCellName,  # I$1    
                                  )

            self.makeNumericCell((thisRow, firstCol),  "=" + expr,
                                 "#,##0.0") 

    #----------------------------------------

    def __fillDataRateAtAveragePileup(self,
                                      topLeft,
                                      triggerRateCellName,
                                      topLeftsUnweightedRates,
                                      weightCells,
                                      ):
        # TODO: could be merged with __fillDataRateAtPileup(..) ?
        assert len(weightCells) == len(topLeftsUnweightedRates)

        firstRow, firstCol = topLeft

        #----------
        # titles
        #----------

        self[(firstRow,    firstCol)] = "data rate [MByte/s]" 
        self[(firstRow + 1,firstCol)] = 'at average pileup'
        self[(firstRow + 2,firstCol)] = '=CONCATENATE("and ", TEXT(%s,"0.0")," kHz trigger rate")' % triggerRateCellName 

        #----------
        # equations
        #----------

        for i in range(len(self.evolutionData)):
            thisRow = firstRow + 3 + i

            parts = []

            for inputRateCell, weightCell in zip(topLeftsUnweightedRates, weightCells):
                parts.append("%s*%s" % (
                        coordToName(inputRateCell[0] + 3 + i, inputRateCell[1]),
                        coordToName(*weightCell)))

            expr = " + ".join(parts)

            self.makeNumericCell((thisRow, firstCol),  "=" + expr, "#,##0.0") 

        
    #----------------------------------------

    def fillMultiPileupProjectionSheet(self, pileups):

        # create a new worksheet
        self.__createWorkSheet()

        #----------
        # average number of vertices and
        # nominal trigger rate
        #----------
        row = self.__makeHeaderCells()

        #----------
        # fill number of pileup lines
        #----------
        
        sumWeightsCell = (row + len(pileups), 8)
        sumWeightsCellName = coordToName(*sumWeightsCell)

        # add the sum of weights cell
        self[(sumWeightsCell[0], sumWeightsCell[1] - 1)] = 'sum of weights'
        self.makeNumericCell(sumWeightsCell, 
                             '=SUM(%s:%s)' % (coordToName(row, sumWeightsCell[1]),
                                              coordToName(row + len(pileups) - 1, sumWeightsCell[1]),
                                              ),
                             "#,##0.000")

        pileupCells = []
        weightCells = []

        for i, pileup in enumerate(pileups):
            self[(row + i, 1)] = 'number of interactions/bx #%d:' % (i+1)
            self.makeNumericCell((row + i, 2), pileup,  "0.0")

            pileupCells.append((row + i, 2))

            self[(row + i, 4)] = 'number of vertices #%d:' % (i+1)
            self[(row + i, 5)] = '=%s*0.7' % coordToName(row + i, 2)
            
            self[(row + i, 7)] = 'weight (arb. units) #%d:' % (i+1)
            self.makeNumericCell((row + i, 8), 1.0 / len(pileups), "#,##0.000")
            
            self[(row + i, 10)] = 'weight (norm.) #%d:' % (i+1)

            weightCells.append((row + i, 11))

            # cell with the normalized weight value
            self.makeNumericCell(weightCells[-1], 
                                 '=%s / %s' % (
                    coordToName(row + i, 8), 
                    sumWeightsCellName), 
                                 "0.00%")

        # advance base row
        row += 2 + len(pileups)

        #----------
        # fill input data (fit results)
        #----------
        self.__fillInputData((row,1))

        topLeftInputData = (row + 3, 4)
        topLeftNumFeds   = (row + 3, 2)

        nextCol = 5 + self.numCoeffs

        #----------
        # when do we reach 4000 GByte/s per fedbuilder ?
        #----------

        if self.numCoeffs <= 3:
            maxDataRateCell = (1, 5)

            # maximum data rate per FED
            if self.groupingName == 'by fedbuilder':
                maxDataRate = 4000 # in MByte/s
                divideByNumFeds = False
            else:
                maxDataRate = 200 # in MByte/s
                divideByNumFeds = True

            # now this is independent on any precalculations
            numCols = self.__fillLimit(topLeft = (row, nextCol), 
                                       topLeftInputData = topLeftInputData,
                                       maxDataRate = maxDataRate,
                                       maxDataRateCell = maxDataRateCell,
                                       triggerRateCellName = self.triggerRateCellName,
                                       divideByNumFeds = divideByNumFeds,
                                       topLeftNumFeds = topLeftNumFeds,
                                       usePileup = True
                                       )

            nextCol += 1 + numCols
        else:
            maxDataRateCell = None

        #--------------------
        # data rate at given trigger rate and given number of pileup
        #--------------------

        topLeftsUnweightedRates = []

        for j, pileupCell in enumerate(pileupCells):

            topLeftsUnweightedRates.append((row, nextCol + 2 * j))

            self.__fillDataRateAtPileup(topLeft = topLeftsUnweightedRates[-1],
                                        topLeftInputData = topLeftInputData,
                                        triggerRateCellName = self.triggerRateCellName,
                                        pileupCellName = coordToName(*pileupCell, rowPrefix = '$', colPrefix = '$')
                                        )

            if maxDataRateCell != None:
                self.__conditionalFormattingForDataRate(topLeftsUnweightedRates[-1],
                                                        maxDataRateCell)


        #--------------------
        # weighted sum of data rates at different pileups
        #--------------------

        if len(pileups) > 1:
            topLeft = (row, nextCol + 2 * len(pileups))
            self.__fillDataRateAtAveragePileup(topLeft = topLeft,
                                               triggerRateCellName = self.triggerRateCellName,
                                               topLeftsUnweightedRates = topLeftsUnweightedRates,
                                               weightCells = weightCells,
                                               )

            if maxDataRateCell != None:
                self.__conditionalFormattingForDataRate(topLeft,
                                                        maxDataRateCell)

    #----------------------------------------

    def makeNumericCell(self, cellName, value, format = None):
        # cellName can either be a string or a pair (row, col)
        if not isinstance(cellName, str):
            # convert from tuple to string
            cellName = coordToName(*cellName)

        self.ws[cellName] = value
        if format != None:
            self.ws.cell(cellName).number_format = format

    #----------------------------------------


#----------------------------------------------------------------------

class SpreadsheetCreator:


    #----------------------------------------

    def __init__(self, subsystemEvolutionData, avgNumVertices = None,
                 triggerRateKHz = 100, pileups = None):
        import openpyxl

        self.subsystemEvolutionData = subsystemEvolutionData
        self.triggerRateKHz = triggerRateKHz
        self.avgNumVertices = avgNumVertices

        if pileups != None:
            assert self.subsystemEvolutionData.has_key('by fedbuilder')

        # create a workbook
        self.wb = openpyxl.Workbook()

        # fill the workbook
        for groupingName, groupingData in self.subsystemEvolutionData.items():
            # make one sheet per grouping
            sheetFiller = SingleGroupSheet(self.wb, groupingName, groupingData, avgNumVertices, triggerRateKHz)
            sheetFiller.fillSheet()

        if pileups != None:
            # add multi-pileup projection sheet for fedbuilder

            groupingName = 'by fedbuilder'
            groupingData = self.subsystemEvolutionData[groupingName]

            sheetFiller = SingleGroupSheet(self.wb, groupingName, groupingData, avgNumVertices, triggerRateKHz, sheetName = "multi pileup " + groupingName)
            sheetFiller.fillMultiPileupProjectionSheet(pileups)

    #----------------------------------------
    def writeToFile(self, outputFname):

        fout = open(outputFname, "w")
        fout.write(self.makeString())
        fout.close()


    #----------------------------------------

    def makeString(self):
        # returns the binary notebook as a string
        # (which could be written to a file etc.)

        import openpyxl

        # see http://stackoverflow.com/a/8714342/288875
        return openpyxl.writer.excel.save_virtual_workbook(self.wb)

    #----------------------------------------


            
#----------------------------------------------------------------------
# main
#----------------------------------------------------------------------
if __name__ == "__main__":

    from optparse import OptionParser
    parser = OptionParser("""

      usage: %prog [options] input_file.pkl

      runs on the given input file(s) and keeps/drops products.
    """
    )

    parser.add_option("--pileups",
                      default = None,
                      type="str",
                      help="comma separated list of pileup values to make projections for",
                      metavar="NP1,NP2")

    (options, ARGV) = parser.parse_args()

    assert len(ARGV) == 1

    if options.pileups != None:
        options.pileups = [ float(x) for x in options.pileups.split(',') ]

    #----------

    tasksFile = ARGV.pop(0)

    import pickle

    tasks = pickle.load(open(tasksFile))['tasks']

    plotDir = os.path.dirname(tasksFile)

    subsystemEvolutionData = GrandUnificationPlot.makeSubsystemEvolutionData(tasks)
    sc = SpreadsheetCreator(subsystemEvolutionData,
                            getAverageNumVerticesFromTasks(tasks),
                            pileups = options.pileups)

    outputFname = os.path.join(plotDir, "evolution.xlsx")

    sc.writeToFile(outputFname)

    print >> sys.stderr,"wrote report to", outputFname

